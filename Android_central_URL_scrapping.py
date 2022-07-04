# -*- coding: utf-8 -*-
"""
Created on Sun Jun 26 18:56:29 2022

@author: jsaum
"""

from bs4 import BeautifulSoup
import requests
import json
import os
import pandas as pd
source = requests.get('https://forums.androidcentral.com/search.php?searchid=45275696')
source.raise_for_status() #method will throw a error if the url is invalid 
source.ok
soup = BeautifulSoup(source.text ,'html.parser')
print(soup)
URLs = soup.find('div',class_= 'searchresults').find_all('li')
len(URLs)
URLs[0].find('div',class_='i').find('h1').find('a')['href']
# scrapping all data 
all_whatsapp_URL_info = []
for URL in URLs:
    try:
        temp = {}
        temp["URLs"] = URL.find('div',class_='i').find('h1').find('a')['href']
        all_whatsapp_URL_info.append(temp)
    except Exception as e:
        print(URL.text)
all_whatsapp_URL_text = json.dumps(all_whatsapp_URL_info, indent=4)
print(all_whatsapp_URL_text)
#for appending data in excel 
from openpyxl import load_workbook
all_whatsapp_URL_info_df = pd.json_normalize(all_whatsapp_URL_info)
book = load_workbook('Android_central_URL.xlsx')

with pd.ExcelWriter('Android_central_URL.xlsx', engine='openpyxl', mode='a') as writer:  
#writer = pd.ExcelWriter('Android_central_URL.xlsx', engine='openpyxl')

    try:
        #df = pd.read_excel( writer, sheet_name='Instagram')
        #if df:
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        all_whatsapp_URL_info_df.to_excel(writer, sheet_name='Growtopia', startrow=writer.sheets['Growtopia'].max_row, index_label=None,index=False,header=False)
        print("Appended to existing sheet! :)")
    except Exception as e: 
        print ("Exception, creating new sheet: %s"%(repr(e)))
        all_whatsapp_URL_info_df.to_excel(writer, sheet_name='Growtopia',index= False,header=False, index_label=None)